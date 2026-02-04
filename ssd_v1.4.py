import streamlit as st
import pandas as pd
import numpy as np
from scipy import interpolate
from scipy.interpolate import interp1d, PchipInterpolator
import matplotlib.pyplot as plt
import os
from io import BytesIO

# FUNÇÕES DE CÁLCULO 

def obter_k_dinamico(area_km2):
    """Calcula coeficiente K dinâmico baseado na área"""
    ha = area_km2 * 100.0
    if ha <= 5:
        return 0.90
    elif ha <= 10:
        return 0.85
    elif ha <= 20:
        return 0.80
    elif ha <= 50:
        return 0.75
    else:
        return 0.70


ordem_meses = {
    'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6,
    'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
}


def wide_to_long_monthly(df, mes_inicial, ano_inicial, mes_final, ano_final):
    """Converte dados de vazão de formato wide para long"""
    columns = ['Ano'] + list(df.columns[1:])
    df.columns = columns
    df = df.iloc[1:].reset_index(drop=True)
    df_long = pd.melt(df, id_vars=['Ano'], value_vars=columns[1:], 
                      var_name='Mês', value_name='Vazão (m³/s)')
    df_long['Ordem_Mês'] = df_long['Mês'].map(ordem_meses)
    df_long['Data'] = pd.to_datetime(
        df_long['Ano'].astype(str) + '-' + df_long['Ordem_Mês'].astype(str) + '-01'
    )
    
    mi, mf = ordem_meses[mes_inicial], ordem_meses[mes_final]
    d_ini = pd.to_datetime(f"{ano_inicial}-{mi}-01")
    d_fim = pd.to_datetime(f"{ano_final}-{mf}-01") + pd.offsets.MonthEnd(0)
    
    return df_long[(df_long['Data'] >= d_ini) & (df_long['Data'] <= d_fim)].sort_values('Data').reset_index(drop=True)


def simular_sistema_n(dfs, params, modo, vazao_conjunta):
    """Função principal de simulação do sistema (mantida intacta)"""
    n_res = len(dfs)
    n_meses = len(dfs[0])
    segundos_mes = 2.592e6

    colunas_init = ['Armazenamento Inicial', 'Armazenamento Final', 'Demanda Solicitada (m³/s)',
                    'Demanda Atendida (m³/s)', 'Racionamento (%)', 'Transferência Recebida (m³/s)',
                    'Transferência Enviada (m³/s)', 'Evaporação (hm³)', 'Vertimento (hm³)', 
                    'Falha', 'Modo Operação']

    for df in dfs:
        for col in colunas_init:
            df[col] = 0.0
        df['Falha'] = 'Não'
        df['Modo Operação'] = 'Normal'

    volumes_atueis = [p['vol_ini'] for p in params]

    for t in range(n_meses):
        demandas_iniciais = []
        racionamentos = []
        nomes_faixas_atuais = []
        prev_volumes_pos_natureza = []

        # Cálculo de racionamento e volumes pós-natureza
        for i in range(n_res):
            p = params[i]
            vol_ini = volumes_atueis[i]
            pct_vol = (vol_ini / p['capacidade']) * 100
            mes_atual = dfs[i].loc[t, 'Mês']

            rac = 0.0
            nome_faixa = "Normal"
            if p['regras_secas']:
                regras = p['regras_secas'].get(mes_atual, [])
                if regras:
                    nome_faixa = "Acima do Teto"
                    for lim, r_val, n_faixa in regras:
                        if pct_vol <= lim:
                            rac = r_val
                            nome_faixa = n_faixa
                            break

            demandas_iniciais.append(p['demanda_nominal'])
            racionamentos.append(rac)
            nomes_faixas_atuais.append(nome_faixa)

            area = p['func_area'](vol_ini)
            kp_dinamico = obter_k_dinamico(area)
            evap_tanque_mm = dfs[i].loc[t, 'Evaporação (m)']
            evap_hm3 = (evap_tanque_mm * kp_dinamico * area) / 1000.0
            afluencia_hm3 = dfs[i].loc[t, 'Vazão (m³/s)'] * (segundos_mes / 1e6)

            dfs[i].loc[t, 'Evaporação (hm³)'] = evap_hm3
            dfs[i].loc[t, 'Afluências (hm³/mês)'] = afluencia_hm3

            vol_pos_natureza = vol_ini + afluencia_hm3 - evap_hm3
            vol_pos_natureza = max(0.0, vol_pos_natureza)
            prev_volumes_pos_natureza.append(vol_pos_natureza)

        total_vol_disponivel = sum(prev_volumes_pos_natureza)

        # Cálculo da demanda total
        rac_inicial_conjunta = racionamentos[0] if len(racionamentos) > 0 else 0.0
        demanda_conjunta_estimada = vazao_conjunta * (1 - rac_inicial_conjunta / 100.0) * (segundos_mes / 1e6)
        
        total_demanda_necessaria = demanda_conjunta_estimada
        for i in range(n_res):
            p = params[i]
            dem_esp_hm3 = p['demanda_nominal'] * (1 - racionamentos[i] / 100.0) * (segundos_mes / 1e6)
            total_demanda_necessaria += dem_esp_hm3

        # Verificação de falha sistêmica
        sistema_em_falha = False
        if modo == "Paralelo" and (total_vol_disponivel < total_demanda_necessaria):
            sistema_em_falha = True
            for i in range(n_res):
                dfs[i].loc[t, 'Falha'] = 'Sim'
                dfs[i].loc[t, 'Modo Operação'] = 'FALHA SISTÊMICA'

        # Lógica de transferência
        demandas_finais = [0.0] * n_res
        transferencias_registradas = [0.0] * n_res
        transferencias_enviadas = [0.0] * n_res

        if not sistema_em_falha:
            responsabilidade_especifica = [p['demanda_nominal'] for p in params]

            if modo == "Paralelo":
                alocacao_conjunta_bruta = [0.0] * n_res
                if n_res > 0:
                    alocacao_conjunta_bruta[0] = vazao_conjunta

                for i in range(n_res - 1):
                    p = params[i]
                    vol_gatilho = p['capacidade'] * (p['gatilho'] / 100)
                    carga_para_mover_bruta = alocacao_conjunta_bruta[i]

                    if carga_para_mover_bruta > 0:
                        if volumes_atueis[i] < vol_gatilho:
                            dem_esp_prox_teorica = responsabilidade_especifica[i + 1] * (1 - racionamentos[i + 1] / 100.0)
                            carga_conj_prox_racionada = carga_para_mover_bruta * (1 - racionamentos[i + 1] / 100.0)
                            demanda_total_prox_hm3 = (dem_esp_prox_teorica + carga_conj_prox_racionada) * (segundos_mes / 1e6)

                            if prev_volumes_pos_natureza[i + 1] >= demanda_total_prox_hm3:
                                alocacao_conjunta_bruta[i] = 0.0
                                alocacao_conjunta_bruta[i + 1] += carga_para_mover_bruta

                for k in range(n_res):
                    dem_esp = responsabilidade_especifica[k] * (1 - racionamentos[k] / 100.0)
                    dem_conj = alocacao_conjunta_bruta[k] * (1 - racionamentos[k] / 100.0)
                    demandas_finais[k] = dem_esp + dem_conj

                for k in range(1, n_res):
                    if alocacao_conjunta_bruta[k] > 0:
                        valor_transf_racionado = alocacao_conjunta_bruta[k] * (1 - racionamentos[k] / 100.0)
                        transferencias_registradas[k] = valor_transf_racionado
                        transferencias_enviadas[k - 1] = valor_transf_racionado

            elif modo == "Série":
                for k in range(n_res):
                    base_demand = demandas_iniciais[k]
                    if k == 0:
                        base_demand += vazao_conjunta
                    demandas_finais[k] = base_demand * (1 - racionamentos[k] / 100.0)

                for i in range(1, n_res):
                    idx_sender = i
                    idx_receiver = i - 1
                    p_receiver = params[idx_receiver]
                    vol_atual_receiver = prev_volumes_pos_natureza[idx_receiver]
                    cap_receiver = p_receiver['capacidade']
                    gatilho_receiver = p_receiver['gatilho'] / 100.0
                    vol_gatilho_A = cap_receiver * gatilho_receiver

                    if vol_atual_receiver < vol_gatilho_A:
                        demanda_alvo_A = demandas_finais[idx_receiver]
                        vol_demanda_hm3 = demanda_alvo_A * (segundos_mes / 1e6)
                        disponivel_sender = prev_volumes_pos_natureza[idx_sender]
                        qtd_transferir_hm3 = min(vol_demanda_hm3, disponivel_sender)

                        prev_volumes_pos_natureza[idx_receiver] += qtd_transferir_hm3
                        prev_volumes_pos_natureza[idx_sender] -= qtd_transferir_hm3

                        fluxo_transf = qtd_transferir_hm3 * (1e6 / segundos_mes)
                        transferencias_registradas[idx_receiver] += fluxo_transf
                        transferencias_enviadas[idx_sender] += fluxo_transf

            else:  # Individual
                for k in range(n_res):
                    base = demandas_iniciais[k]
                    if k == 0:
                        base += vazao_conjunta
                    demandas_finais[k] = base * (1 - racionamentos[k] / 100.0)
        else:
            for k in range(n_res):
                base = demandas_iniciais[k]
                if modo == "Paralelo" and k == 0:
                    base += vazao_conjunta
                demandas_finais[k] = base * (1 - racionamentos[k] / 100.0)

        # Atualização de volumes
        for i in range(n_res):
            p = params[i]
            df = dfs[i]
            vol_ini = volumes_atueis[i]
            demanda_ms = demandas_finais[i]
            demanda_hm3 = demanda_ms * (segundos_mes / 1e6)

            df.loc[t, 'Demanda Solicitada (m³/s)'] = demandas_iniciais[i]
            if i == 0:
                df.loc[t, 'Demanda Solicitada (m³/s)'] += vazao_conjunta

            df.loc[t, 'Transferência Recebida (m³/s)'] = transferencias_registradas[i]
            df.loc[t, 'Transferência Enviada (m³/s)'] = transferencias_enviadas[i]
            df.loc[t, 'Armazenamento Inicial'] = vol_ini
            df.loc[t, 'Racionamento (%)'] = racionamentos[i]
            
            if not sistema_em_falha:
                df.loc[t, 'Modo Operação'] = nomes_faixas_atuais[i]

            if modo == "Série":
                vol_disp = prev_volumes_pos_natureza[i]
                evap_hm3 = df.loc[t, 'Evaporação (hm³)']
                afluencia_hm3 = df.loc[t, 'Afluências (hm³/mês)']
            else:
                evap_hm3 = df.loc[t, 'Evaporação (hm³)']
                afluencia_hm3 = df.loc[t, 'Afluências (hm³/mês)']
                vol_disp = vol_ini + afluencia_hm3 - evap_hm3

            demanda_atendida_real_hm3 = 0.0

            if sistema_em_falha:
                demanda_atendida_real_hm3 = max(0, min(vol_disp, demanda_hm3))
            else:
                if vol_disp < demanda_hm3:
                    demanda_atendida_real_hm3 = max(0, vol_disp)
                    df.loc[t, 'Falha'] = 'Sim'
                else:
                    demanda_atendida_real_hm3 = demanda_hm3

            real_flow_m3s = demanda_atendida_real_hm3 * (1e6 / segundos_mes)
            df.loc[t, 'Demanda Atendida (m³/s)'] = real_flow_m3s

            vol_final = vol_disp - demanda_atendida_real_hm3
            vertimento = 0.0
            
            if vol_final > p['capacidade']:
                vertimento = vol_final - p['capacidade']
                vol_final = p['capacidade']
            if vol_final < 0:
                vol_final = 0.0

            df.loc[t, 'Vertimento (hm³)'] = vertimento
            df.loc[t, 'Armazenamento Final'] = vol_final
            volumes_atueis[i] = vol_final

    return dfs

# FUNÇÕES DE CARREGAMENTO DE DADOS 

@st.cache_data
def load_data():
    """Carrega todos os dados necessários dos arquivos Excel locais"""
    try:
        base_path = os.path.abspath(".")
        
        caminho_dados = os.path.join(base_path, "Dados_açudes_monitorados.xlsx")
        caminho_vazoes = os.path.join(base_path, "Series_Vazoes_Totais_m3s_Oficiais_Q90_Acudes.xlsx")
        
        if not os.path.exists(caminho_dados):
            st.error(f"❌ Arquivo não encontrado: {caminho_dados}")
            return None
        
        if not os.path.exists(caminho_vazoes):
            st.error(f"❌ Arquivo não encontrado: {caminho_vazoes}")
            return None
        
        # Carregar dados físicos
        xls_dados = pd.ExcelFile(caminho_dados)
        acudes_original = pd.read_excel(xls_dados, "acudes_original")
        
        if "COD" in acudes_original.columns:
            acudes_original["COD"] = acudes_original["COD"].astype(str).str.replace(r'\.0$', '', regex=True)
        
        if "CAPAC (m³)" in acudes_original.columns:
            acudes_original["CAPAC (m³)"] = acudes_original["CAPAC (m³)"] / 1e6
        
        cav = pd.read_excel(xls_dados, "cav")
        if "COD" in cav.columns:
            cav["COD"] = cav["COD"].astype(str).str.replace(r'\.0$', '', regex=True)
        
        if "VOLUME (m³)" in cav.columns and cav["VOLUME (m³)"].mean() > 10000:
            cav["VOLUME (m³)"] /= 1e6
        
        evaporacao = pd.read_excel(xls_dados, "evaporação")
        if "COD" in evaporacao.columns:
            evaporacao["COD"] = evaporacao["COD"].astype(str).str.replace(r'\.0$', '', regex=True)
        
        plano_secas = pd.read_excel(xls_dados, "plano_secas")
        if "COD" in plano_secas.columns:
            plano_secas["COD"] = plano_secas["COD"].astype(str).str.replace(r'\.0$', '', regex=True)
        
        # Carregar presets de hidrossistemas com modo de operação
        presets = {"--- Selecione um Sistema ---": {'modo': None, 'reservatorios': []}}
        try:
            df_hidro = pd.read_excel(xls_dados, "hidrossistemas", header=None)
            
            # Pular a primeira linha se for cabeçalho
            start_row = 1 if df_hidro.iloc[0, 0] == 'hidrossistema' else 0
            
            for index in range(start_row, len(df_hidro)):
                row = df_hidro.iloc[index]
                nome_sistema = row.iloc[0]
                
                if pd.isna(nome_sistema) or str(nome_sistema).strip() == "":
                    continue
                
                # Pegar modo de operação da segunda coluna
                modo_operacao = str(row.iloc[1]).strip().lower() if not pd.isna(row.iloc[1]) else "individual"
                
                # Normalizar o modo
                if 'paral' in modo_operacao or 'paralel' in modo_operacao:
                    modo_operacao = "Paralelo"
                elif 'ser' in modo_operacao or 'série' in modo_operacao:
                    modo_operacao = "Série"
                else:
                    modo_operacao = "Individual"
                
                # Pegar códigos dos reservatórios (a partir da terceira coluna)
                valores_linha = row.iloc[2:].dropna()
                lista_res_limpa = []
                for val in valores_linha:
                    try:
                        cod_str = str(int(float(val)))
                    except:
                        cod_str = str(val).strip()
                    lista_res_limpa.append(cod_str)
                
                if lista_res_limpa:
                    presets[str(nome_sistema)] = {
                        'modo': modo_operacao,
                        'reservatorios': lista_res_limpa
                    }
        except Exception as e:
            print(f"Erro ao carregar presets: {e}")
        
        # Retornar apenas dados serializáveis (sem xls_vazoes)
        return {
            'acudes_original': acudes_original,
            'cav': cav,
            'evaporacao': evaporacao,
            'plano_secas': plano_secas,
            'presets': presets,
            'caminho_dados': caminho_dados,
            'caminho_vazoes': caminho_vazoes  # Guardar apenas o caminho
        }
    
    except Exception as e:
        st.error(f"❌ Erro ao carregar dados: {str(e)}")
        return None


def carregar_vazao_reservatorio(caminho_vazoes, nome_reservatorio):
    """Carrega dados de vazão para um reservatório específico"""
    try:
        df_vazao = pd.read_excel(caminho_vazoes, sheet_name=nome_reservatorio)
        return df_vazao
    except Exception as e:
        st.error(f"❌ Erro ao carregar vazões do reservatório {nome_reservatorio}: {str(e)}")
        return None



# VISUALIZAÇÃO E EDIÇÃO DO PLANO DE SECAS


def visualizar_plano_secas(plano_secas_df, reservatorios):
    """Gera figura com as curvas de racionamento para cada reservatório do hidrossistema"""
    meses_labels = list(ordem_meses.keys())

    cores_faixas = {
        'Seca Severa': '#E53935',
        'Seca':        '#FB8C00',
        'Alerta':      '#FDD835',
        'Normal':      '#43A047',
    }

    n = len(reservatorios)
    fig, axes = plt.subplots(n, 1, figsize=(13, 4.2 * n), dpi=100)
    if n == 1:
        axes = [axes]

    for idx, res in enumerate(reservatorios):
        ax = axes[idx]
        cod = str(res['cod'])

        plano_res = plano_secas_df[plano_secas_df['COD'].astype(str) == cod].copy()

        ax.set_title(f"{res['nome']}  (COD {cod})", fontsize=13, fontweight='bold', pad=10)
        ax.set_ylabel("Volume (%)", fontsize=10)
        ax.set_ylim(0, 108)
        ax.set_xlim(-0.4, 11.4)
        ax.set_xticks(range(12))
        ax.set_xticklabels(meses_labels, fontsize=9)
        ax.grid(True, linestyle='--', alpha=0.4)

        if plano_res.empty:
            ax.text(0.5, 0.5, "Sem Plano de Secas definido",
                    ha='center', va='center', transform=ax.transAxes,
                    fontsize=12, color='#888888', style='italic')
            continue

        # Ordenar faixas do menor limite médio para o maior
        plano_res['_media'] = plano_res[meses_labels].mean(axis=1)
        plano_res = plano_res.sort_values('_media', ascending=True).reset_index(drop=True)

        y_base = np.zeros(12)

        for _, row in plano_res.iterrows():
            faixa = row['Faixa']
            rac   = row['Racionamento (%)']
            y_topo = row[meses_labels].values.astype(float)

            cor = cores_faixas.get(faixa, '#90CAF9')

            ax.fill_between(range(12), y_base, y_topo,
                            color=cor, alpha=0.75, label=f"{faixa}  ({rac:.0f}% rac.)")
            ax.plot(range(12), y_topo, color='white', linewidth=0.8, alpha=0.6)

            y_base = y_topo.copy()


        if np.mean(y_base) < 99:
            ax.fill_between(range(12), y_base, 100,
                            color=cores_faixas['Normal'], alpha=0.75, label='Normal  (0% rac.)')

        ax.legend(loc='lower right', fontsize=8.5, framealpha=0.95)

    fig.tight_layout(pad=2.2)
    return fig


def _plano_efetivo(data):
    """Monta o plano completo: original do xlsx + edições pendentes + planos custom (sessão)."""
    base = data['plano_secas'].copy()
    meses_labels = list(ordem_meses.keys())
    cols = ['Faixa', 'Racionamento (%)'] + meses_labels


    for cod, df_ed in st.session_state.get('plano_secas_editado', {}).items():
        base = base[base['COD'].astype(str) != cod]                  
        df_ins = df_ed[cols].copy()
        df_ins.insert(0, 'COD', cod)
        base = pd.concat([base, df_ins], ignore_index=True)


    for cod, df_ed in st.session_state.get('plano_secas_custom', {}).items():
        df_ins = df_ed[cols].copy()
        df_ins.insert(0, 'COD', cod)
        base = pd.concat([base, df_ins], ignore_index=True)

    return base


def editar_plano_secas(plano_secas_df, reservatorios, caminho_dados):
    """Editor com adicionar / remover faixas.
    CODs que existem no xlsx → edições vão para 'plano_secas_editado' e podem ser salvas.
    CODs que NÃO existem no xlsx (sistema custom) → vão para 'plano_secas_custom', só sessão.
    """
    meses_labels  = list(ordem_meses.keys())
    cols_edit     = ['Faixa', 'Racionamento (%)'] + meses_labels

    # garantir chaves no session_state
    if 'plano_secas_editado' not in st.session_state:
        st.session_state['plano_secas_editado'] = {}
    if 'plano_secas_custom' not in st.session_state:
        st.session_state['plano_secas_custom']  = {}

    codigos_no_xlsx = set(plano_secas_df['COD'].astype(str).unique())

    for res in reservatorios:
        cod       = str(res['cod'])
        is_custom = cod not in codigos_no_xlsx   
        state_key = 'plano_secas_custom' if is_custom else 'plano_secas_editado'


        if cod in st.session_state[state_key]:
            plano_res = st.session_state[state_key][cod].copy()
        else:
            sub = plano_secas_df[plano_secas_df['COD'].astype(str) == cod][cols_edit]
            plano_res = sub.reset_index(drop=True) if not sub.empty else pd.DataFrame(columns=cols_edit)

       
        st.subheader(f"🌊 {res['nome']}  (COD {cod})")

        if st.button("＋ Adicionar Faixa", key=f"add_faixa_{cod}"):
            nova = pd.DataFrame([{
                'Faixa': 'Nova Faixa',
                'Racionamento (%)': 0.0,
                **{m: 50.0 for m in meses_labels}
            }])
            plano_res = pd.concat([plano_res, nova], ignore_index=True)
            st.session_state[state_key][cod] = plano_res.copy()
            st.rerun()

       
        if len(plano_res) > 1:
            col_sel, col_btn = st.columns([3, 1])
            with col_sel:
                faixa_rem = st.selectbox(
                    "Remover faixa:",
                    plano_res['Faixa'].tolist(),
                    key=f"sel_rem_{cod}",
                    label_visibility="collapsed"
                )
            with col_btn:
                if st.button("− Remover", key=f"btn_rem_{cod}"):
                    idx = plano_res[plano_res['Faixa'] == faixa_rem].index[0]
                    plano_res = plano_res.drop(index=idx).reset_index(drop=True)
                    st.session_state[state_key][cod] = plano_res.copy()
                    st.rerun()

        # ---- tabela editável ----
        if plano_res.empty:
            st.warning("Sem faixas. Clique '＋ Adicionar Faixa' para começar.")
        else:
            edited = st.data_editor(
                plano_res,
                key=f"editor_plano_{cod}",
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Faixa': st.column_config.TextColumn(),                   # editável
                    'Racionamento (%)': st.column_config.NumberColumn(
                        min_value=0.0, max_value=100.0, step=1.0, format="%.1f"),
                    **{m: st.column_config.NumberColumn(
                        min_value=0.0, max_value=100.0, step=1.0, format="%.0f") for m in meses_labels}
                }
            )
            # persistir mudanças imediatas no session_state
            st.session_state[state_key][cod] = edited.copy()

        st.markdown("---")

    # Salvar no .xlsx
    if st.session_state.get('plano_secas_editado'):
        st.markdown("### 💾 Salvar Alterações no Arquivo")

        if st.button("💾 Salvar Planos de Secas no Excel", type="primary", use_container_width=True):
            try:
                import openpyxl

                wb     = openpyxl.load_workbook(caminho_dados)
                ws     = wb['plano_secas']
                header = [c.value for c in ws[1]]

                cod_col   = header.index('COD') + 1
                faixa_col = header.index('Faixa') + 1
                rac_col   = header.index('Racionamento (%)') + 1
                mes_cols  = {m: header.index(m) + 1 for m in meses_labels}

                rows_to_del = []
                for row_idx in range(2, ws.max_row + 1):
                    cv = str(ws.cell(row=row_idx, column=cod_col).value).replace('.0', '')
                    if cv in st.session_state['plano_secas_editado']:
                        rows_to_del.append(row_idx)
                for ri in sorted(rows_to_del, reverse=True):
                    ws.delete_rows(ri)

      
                for cod_val, df_ed in st.session_state['plano_secas_editado'].items():
                    for _, nr in df_ed.iterrows():
                        nova_row = ws.max_row + 1
                        ws.cell(row=nova_row, column=cod_col,
                                value=int(cod_val) if cod_val.isdigit() else cod_val)
                        ws.cell(row=nova_row, column=faixa_col, value=str(nr['Faixa']))
                        ws.cell(row=nova_row, column=rac_col,   value=float(nr['Racionamento (%)']))
                        for m in meses_labels:
                            ws.cell(row=nova_row, column=mes_cols[m], value=float(nr[m]))

                wb.save(caminho_dados)

                st.session_state['plano_secas_editado'] = {}
                load_data.clear()
                st.success("✅ Planos de secas salvos com sucesso!")
                st.rerun()

            except Exception as e:
                st.error(f"❌ Erro ao salvar: {str(e)}")

    # info sobre planos custom
    if st.session_state.get('plano_secas_custom'):
        st.info("ℹ️ Planos marcados como *novo (sessão)* existem apenas durante esta sessão e são usados diretamente pela simulação — não são gravados no Excel.")



# INTERFACE

def main():
    st.set_page_config(
        page_title="Simulador Hidrológico",
        page_icon="icone.ico",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    st.title("Sistema de Suporte a Decisão")
    
    # Carregar dados
    data = load_data()
    
    if data is None:
        st.error("⚠️ Não foi possível carregar os dados. Verifique se os arquivos Excel estão no diretório correto.")
        return
    
    st.success("✅ Dados carregados com sucesso!")
    
    # ---- session_state defaults ----
    defaults = {
        'reservatorios_selecionados': [],
        'resultados_simulacao':       None,
        'mostrar_plano_secas':        False,
        'mostrar_editor_secas':       False,
        'plano_secas_editado':        {},
        'plano_secas_custom':         {},
        'modo_criacao':               'preset',   # 'preset' | 'custom'
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # Criar abas
    tab1, tab2, tab3 = st.tabs(["⚙️ Configuração do Sistema", "📊 Gráficos", "📈 Análise de Garantia"])
    
    # CONFIGURAÇÃO DO SISTEMA

    with tab1:
        st.header("Configuração do Hidrossistema")

        # ---- badge: sistema atual ----
        if st.session_state.get('ultimo_preset_carregado'):
            modo_info = ""
            if st.session_state.get('modo_operacao_predefinido'):
                modo_info = f" | 🔧 Modo: **{st.session_state['modo_operacao_predefinido']}**"
            st.info(f"📌 **Sistema Atual:** {st.session_state['ultimo_preset_carregado']} "
                    f"({len(st.session_state.reservatorios_selecionados)} reservatório(s)){modo_info}")
        elif st.session_state.reservatorios_selecionados and st.session_state['modo_criacao'] == 'custom':
            st.info(f"📌 **Sistema Custom** — {len(st.session_state.reservatorios_selecionados)} reservatório(s) | 🔧 Modo: livre")

        col_t1, col_t2 = st.columns(2)
        with col_t1:
            btn_preset_active = (st.session_state['modo_criacao'] == 'preset')
            if st.button("📦 Usar Sistema Predefinido",
                         type="primary" if btn_preset_active else "secondary",
                         use_container_width=True):
                if not btn_preset_active:
                    # reset tudo e mudar modo
                    st.session_state.reservatorios_selecionados = []
                    st.session_state.resultados_simulacao       = None
                    st.session_state['mostrar_plano_secas']     = False
                    st.session_state['mostrar_editor_secas']    = False
                    st.session_state['plano_secas_editado']     = {}
                    st.session_state['plano_secas_custom']      = {}
                    st.session_state.pop('ultimo_preset_carregado',   None)
                    st.session_state.pop('modo_operacao_predefinido', None)
                    # limpar estado interno dos widgets de parâmetros
                    st.session_state.pop('modo_op',     None)
                    st.session_state.pop('vazao_conj',  None)
                    st.session_state['modo_criacao'] = 'preset'
                    st.rerun()

        with col_t2:
            btn_custom_active = (st.session_state['modo_criacao'] == 'custom')
            if st.button("🛠️ Simular Outro Hidrossistema",
                         type="primary" if btn_custom_active else "secondary",
                         use_container_width=True):
                if not btn_custom_active:
                    st.session_state.reservatorios_selecionados = []
                    st.session_state.resultados_simulacao       = None
                    st.session_state['mostrar_plano_secas']     = False
                    st.session_state['mostrar_editor_secas']    = False
                    st.session_state['plano_secas_editado']     = {}
                    st.session_state['plano_secas_custom']      = {}
                    st.session_state.pop('ultimo_preset_carregado',   None)
                    st.session_state.pop('modo_operacao_predefinido', None)
                    # limpar estado interno dos widgets de parâmetros
                    st.session_state.pop('modo_op',     None)
                    st.session_state.pop('vazao_conj',  None)
                    st.session_state['modo_criacao'] = 'custom'
                    st.rerun()

        st.markdown("---")


        # SISTEMA PREDEFINIDO

        if st.session_state['modo_criacao'] == 'preset':

            if st.session_state.resultados_simulacao is not None:
                st.warning("⚠️ **Atenção:** Ao carregar um novo sistema, todas as simulações anteriores serão apagadas da memória.")

            col1, col2 = st.columns([2, 1])
            with col1:
                preset_selecionado = st.selectbox(
                    "Selecione um Sistema Predefinido:",
                    options=list(data['presets'].keys()),
                    key="preset_select"
                )
            with col2:
                if st.button("🔄 Carregar Sistema", type="primary"):
                    if preset_selecionado != "--- Selecione um Sistema ---":
                        # reset
                        st.session_state.reservatorios_selecionados = []
                        st.session_state.resultados_simulacao       = None
                        st.session_state['mostrar_plano_secas']     = False
                        st.session_state['mostrar_editor_secas']    = False
                        st.session_state['plano_secas_editado']     = {}
                        st.session_state['plano_secas_custom']      = {}
                        st.session_state.pop('ultimo_preset_carregado', None)
                        # limpar estado interno dos widgets de parâmetros
                        st.session_state.pop('modo_op',     None)
                        st.session_state.pop('vazao_conj',  None)

                        preset_data      = data['presets'][preset_selecionado]
                        modo_predefinido = preset_data['modo']

                        for cod in preset_data['reservatorios']:
                            row = data['acudes_original'][
                                data['acudes_original']["COD"].astype(str) == str(cod)
                            ]
                            if not row.empty:
                                d = row.iloc[0]
                                st.session_state.reservatorios_selecionados.append({
                                    'nome':        d["CORPO"],
                                    'cod':         d["COD"],
                                    'capacidade':  d["CAPAC (m³)"],
                                    'est_evap':    d["Est. Evap."],
                                    'vol_inicial': d["CAPAC (m³)"],
                                    'demanda':     0.0,
                                    'gatilho':     10.0
                                })

                        st.session_state['ultimo_preset_carregado']   = preset_selecionado
                        st.session_state['modo_operacao_predefinido'] = modo_predefinido
                        st.rerun()

            # ---- tabela de reservatórios ----
            if st.session_state.reservatorios_selecionados:
                df_edit = pd.DataFrame(st.session_state.reservatorios_selecionados)
                df_edit['ordem'] = range(1, len(df_edit) + 1)
                df_display = df_edit[['ordem', 'nome', 'capacidade', 'vol_inicial', 'demanda', 'gatilho']]
                df_display.columns = ['Ordem', 'Nome', 'Capacidade (hm³)', 'Vol. Inicial (hm³)',
                                      'Demanda Específica (m³/s)', 'Gatilho (%)']
                edited_df = st.data_editor(
                    df_display, use_container_width=True, hide_index=True,
                    disabled=['Ordem', 'Nome', 'Capacidade (hm³)'],
                    key="editor_reservatorios"
                )
                for idx, row in edited_df.iterrows():
                    st.session_state.reservatorios_selecionados[idx]['vol_inicial'] = row['Vol. Inicial (hm³)']
                    st.session_state.reservatorios_selecionados[idx]['demanda']     = row['Demanda Específica (m³/s)']
                    st.session_state.reservatorios_selecionados[idx]['gatilho']     = row['Gatilho (%)']

                if st.button("🗑️ Limpar Todos os Reservatórios"):
                    st.session_state.reservatorios_selecionados = []
                    st.session_state.resultados_simulacao       = None
                    st.session_state.pop('modo_op',    None)
                    st.session_state.pop('vazao_conj', None)
                    st.rerun()
            else:
                st.info("ℹ️ Nenhum sistema carregado. Selecione um preset acima.")


        # SISTEMA CUSTOMIZADO

        else:   # modo_criacao == 'custom'
            acudes = data['acudes_original']
            lista_nomes = sorted(acudes['CORPO'].tolist())

            st.subheader("🔍 Adicionar Reservatório ao Sistema")
            col_busca, col_add = st.columns([3, 1])
            with col_busca:
                nome_sel = st.selectbox(
                    "Busque um açude:",
                    ["— selecione —"] + lista_nomes,
                    key="busca_acude",
                    label_visibility="collapsed"
                )
            with col_add:
                if st.button("＋ Adicionar", key="btn_add_acude"):
                    if nome_sel == "— selecione —":
                        st.warning("Selecione um açude primeiro.")
                    elif any(r['nome'] == nome_sel for r in st.session_state.reservatorios_selecionados):
                        st.warning(f"'{nome_sel}' já está no sistema.")
                    else:
                        row = acudes[acudes['CORPO'] == nome_sel].iloc[0]
                        st.session_state.reservatorios_selecionados.append({
                            'nome':        row['CORPO'],
                            'cod':         row['COD'],
                            'capacidade':  row['CAPAC (m³)'],
                            'est_evap':    row['Est. Evap.'],
                            'vol_inicial': row['CAPAC (m³)'],
                            'demanda':     0.0,
                            'gatilho':     10.0
                        })
                        st.rerun()

         
            if st.session_state.reservatorios_selecionados:
                df_edit = pd.DataFrame(st.session_state.reservatorios_selecionados)
                df_edit['ordem'] = range(1, len(df_edit) + 1)
                df_display = df_edit[['ordem', 'nome', 'capacidade', 'vol_inicial', 'demanda', 'gatilho']]
                df_display.columns = ['Ordem', 'Nome', 'Capacidade (hm³)', 'Vol. Inicial (hm³)',
                                      'Demanda (m³/s)', 'Gatilho (%)']
                edited_df = st.data_editor(
                    df_display, use_container_width=True, hide_index=True,
                    disabled=['Ordem', 'Nome', 'Capacidade (hm³)'],
                    key="editor_reserv_custom"
                )
                for idx, row in edited_df.iterrows():
                    st.session_state.reservatorios_selecionados[idx]['vol_inicial'] = row['Vol. Inicial (hm³)']
                    st.session_state.reservatorios_selecionados[idx]['demanda']     = row['Demanda (m³/s)']
                    st.session_state.reservatorios_selecionados[idx]['gatilho']     = row['Gatilho (%)']

                # remover reservatório individual
                nomes_atual = [r['nome'] for r in st.session_state.reservatorios_selecionados]
                col_r1, col_r2 = st.columns([3, 1])
                with col_r1:
                    nome_rem = st.selectbox("Remover:", nomes_atual,
                                            key="sel_rem_res", label_visibility="collapsed")
                with col_r2:
                    if st.button("− Remover", key="btn_rem_res"):
                        st.session_state.reservatorios_selecionados = [
                            r for r in st.session_state.reservatorios_selecionados if r['nome'] != nome_rem
                        ]
                        st.rerun()

                if st.button("🗑️ Limpar Todos"):
                    st.session_state.reservatorios_selecionados = []
                    st.session_state.resultados_simulacao       = None
                    st.session_state['plano_secas_custom']      = {}
                    st.session_state.pop('modo_op',    None)
                    st.session_state.pop('vazao_conj', None)
                    st.rerun()
            else:
                st.info("ℹ️ Busque e adicione açudes acima para montar seu sistema.")

        # BOTÕES VER / EDITAR PLANO DE SECAS  
        
        if st.session_state.reservatorios_selecionados:
            st.markdown("")
            col_ver, col_edit = st.columns(2)

            with col_ver:
                if st.button("📊 Ver Níveis Meta", use_container_width=True):
                    st.session_state['mostrar_plano_secas']  = not st.session_state.get('mostrar_plano_secas', False)
                    st.session_state['mostrar_editor_secas'] = False
                    st.rerun()

            with col_edit:
                if st.button("✏️ Editar Planos de Secas", use_container_width=True):
                    st.session_state['mostrar_editor_secas'] = not st.session_state.get('mostrar_editor_secas', False)
                    st.session_state['mostrar_plano_secas']  = False
                    st.rerun()

            # Visualização das curvas
            if st.session_state.get('mostrar_plano_secas', False):
                st.markdown("#### 📈 Curvas de Racionamento")
                fig = visualizar_plano_secas(_plano_efetivo(data),
                                            st.session_state.reservatorios_selecionados)
                st.pyplot(fig)
                plt.close(fig)

            # Editor co plano
            if st.session_state.get('mostrar_editor_secas', False):
                st.markdown("#### ✏️ Editar Planos de Secas")
                editar_plano_secas(
                    data['plano_secas'],
                    st.session_state.reservatorios_selecionados,
                    data['caminho_dados']
                )

        st.markdown("---")
        
        # Parâmetros da Simulação
        st.subheader("Parâmetros da Simulação")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            mes_inicial = st.selectbox("Mês Inicial:", list(ordem_meses.keys()), key="mes_ini")
        
        with col2:
            ano_inicial = st.number_input("Ano Inicial:", min_value=1911, max_value=2016, 
                                         value=1911, key="ano_ini")
        
        with col3:
            mes_final = st.selectbox("Mês Final:", list(ordem_meses.keys()), 
                                    index=11, key="mes_fim")
        
        with col4:
            ano_final = st.number_input("Ano Final:", min_value=1912, max_value=2017, 
                                       value=2017, key="ano_fim")
        
        col5, col6 = st.columns(2)
        
        with col5:
            n_res = len(st.session_state.reservatorios_selecionados)
            modo_options = ["Individual"]
            if n_res >= 2:
                modo_options = ["Paralelo", "Individual", "Série"]

            if st.session_state.get('modo_operacao_predefinido'):
                modo_predefinido = st.session_state['modo_operacao_predefinido']
                modo = st.selectbox(
                    "Modo de Operação:",
                    [modo_predefinido],
                    key="modo_op",
                    disabled=True,
                    help=f"🔒 Modo travado pelo sistema '{st.session_state.get('ultimo_preset_carregado', '')}'"
                )
            else:
                modo = st.selectbox("Modo de Operação:", modo_options, key="modo_op")
        
        with col6:
            vazao_conjunta_disabled = (modo == "Individual")
            vazao_conjunta = st.number_input(
                "Vazão Conjunta (m³/s):",
                min_value=0.0,
                value=0.0,
                step=0.1,
                disabled=vazao_conjunta_disabled,
                key="vazao_conj"
            )

        # Forçar zero quando o modo não usa vazão conjunta.
        if modo == "Individual":
            vazao_conjunta = 0.0
        
        st.markdown("---")
        
        # Botão de processamento
        if st.button("🚀 PROCESSAR SIMULAÇÃO", type="primary", use_container_width=True):
            if not st.session_state.reservatorios_selecionados:
                st.error("⚠️ Adicione reservatórios antes de processar!")
            else:
                with st.spinner("Processando simulação..."):
                    try:
                        processar_simulacao(data, modo, vazao_conjunta, 
                                          mes_inicial, ano_inicial, mes_final, ano_final)
                        st.success("✅ Simulação concluída com sucesso!")
                    except Exception as e:
                        st.error(f"❌ Erro na simulação: {str(e)}")
    
    # RESULTADOS

    with tab2:
        st.header("Resultados da Simulação")
        
        if st.session_state.resultados_simulacao is None:
            st.info("ℹ️ Execute a simulação na aba 'Configuração do Sistema' para visualizar os resultados.")
        else:
            exibir_resultados()
    
    # ANÁLISE DE GARANTIA
    
    with tab3:
        st.header("Análise de Permanência e Garantia")
        
        if st.session_state.resultados_simulacao is None:
            st.info("ℹ️ Execute a simulação na aba 'Configuração do Sistema' para visualizar a análise.")
        else:
            exibir_analise_garantia()


def processar_simulacao(data, modo, vazao_conjunta, mes_ini, ano_ini, mes_fim, ano_fim):
    """Processa a simulação hidrológica"""
    segundos_mes = 2.592e6

    plano_eff = _plano_efetivo(data)
    
    lista_dfs_input = []
    lista_params = []
    
    for res in st.session_state.reservatorios_selecionados:
        # Carregar vazões usando o caminho armazenado
        df_vazao = carregar_vazao_reservatorio(data['caminho_vazoes'], res['nome'])
        
        if df_vazao is None:
            st.error(f"❌ Erro ao carregar vazões do reservatório {res['nome']}")
            return
        
        df_long = wide_to_long_monthly(df_vazao, mes_ini, ano_ini, mes_fim, ano_fim)
        
        # Evaporação
        cod_evap = str(res['est_evap']).replace('.0', '').strip()
        evap_row = data['evaporacao'][data['evaporacao']["COD"] == cod_evap]
        
        if not evap_row.empty:
            df_long["Evaporação (m)"] = df_long["Mês"].map(evap_row.iloc[0][list(ordem_meses.keys())])
        else:
            df_long["Evaporação (m)"] = 0.0
        
        # Função de Interpolação
        cod_acude = str(res['cod'])
        cav_data = data['cav'][data['cav']["COD"] == cod_acude]
        
        if len(cav_data) < 2:
            x_vol = cav_data["VOLUME (m³)"].values
            y_area = cav_data["AREA (km²)"].values
            x_vol = np.append(x_vol, 0.0)
            y_area = np.append(y_area, 0.0)
            unique_indices = np.unique(x_vol, return_index=True)[1]
            x_vol = x_vol[unique_indices]
            y_area = y_area[unique_indices]
            
            if len(x_vol) < 2:
                func_interp = lambda v: 0.0
            else:
                func_interp = interpolate.interp1d(x_vol, y_area, fill_value="extrapolate")
        else:
            cav_sorted = cav_data.sort_values("VOLUME (m³)")
            x_vol = cav_sorted["VOLUME (m³)"].values
            y_area = cav_sorted["AREA (km²)"].values
            try:
                func_interp = PchipInterpolator(x_vol, y_area, extrapolate=True)
            except:
                func_interp = interpolate.interp1d(x_vol, y_area, fill_value="extrapolate")
        
        # Carregar plano de secas do plano efetivo (inclui edições + custom)
        plano = plano_eff[plano_eff['COD'].astype(str) == cod_acude].copy()
        regras_mes = {}
        
        if not plano.empty:
            for m in ordem_meses.keys():
                regras = []
                for _, row in plano.iterrows():
                    regras.append((row[m], row['Racionamento (%)'], row['Faixa']))
                regras.sort(key=lambda x: x[0])
                regras_mes[m] = regras
        
        lista_dfs_input.append(df_long)
        lista_params.append({
            'func_area': func_interp,
            'regras_secas': regras_mes,
            'capacidade': res['capacidade'],
            'vol_ini': res['vol_inicial'],
            'demanda_nominal': res['demanda'],
            'gatilho': res['gatilho']
        })
    
    # Executar simulação
    dfs_resultados = simular_sistema_n(lista_dfs_input, lista_params, modo, vazao_conjunta)
    
    # Salvar resultados no session_state
    st.session_state.resultados_simulacao = {
        'dfs': dfs_resultados,
        'params': lista_params,
        'modo': modo,
        'vazao_conjunta': vazao_conjunta
    }


def exibir_resultados():
    """Exibe os gráficos de resultados da simulação"""
    resultados = st.session_state.resultados_simulacao
    dfs = resultados['dfs']
    n = len(dfs)
    
    # Criar figura com subplots
    fig, axes = plt.subplots(n, 1, figsize=(15, 4 * n), dpi=100)
    
    if n == 1:
        axes = [axes]
    
    for i, (df, ax) in enumerate(zip(dfs, axes)):
        nome = st.session_state.reservatorios_selecionados[i]['nome']
        cap = st.session_state.reservatorios_selecionados[i]['capacidade']
        
        datas = df['Data']
        vol_pct = (df['Armazenamento Final'] / cap) * 100
        
        ax.plot(datas, vol_pct, label='Volume (%)', color='blue', linewidth=2)
        ax.fill_between(datas, 0, vol_pct, alpha=0.3, color='blue')
        
        # Marcar falhas
        falhas = df[df['Falha'] == 'Sim']
        if not falhas.empty:
            ax.scatter(falhas['Data'], [0] * len(falhas), color='red', 
                      marker='x', s=100, label='Falha', zorder=5)
        
        # Marcar transferências
        transf_recebida = df[df['Transferência Recebida (m³/s)'] > 0]
        if not transf_recebida.empty:
            ax.scatter(transf_recebida['Data'], 
                      (transf_recebida['Armazenamento Final'] / cap) * 100,
                      color='orange', marker='v', s=80, label='Transferência', zorder=5)
        
        ax.set_title(f"{nome} (Capacidade: {cap:.1f} hm³)", fontsize=12, fontweight='bold')
        ax.set_ylabel("Volume (%)", fontsize=10)
        ax.set_ylim(0, 110)
        ax.grid(True, linestyle='--', alpha=0.5)
        ax.legend(loc='upper right')
    
    plt.tight_layout()
    st.pyplot(fig)
    
    # Botão de download
    st.markdown("---")
    
    buffer = criar_excel_resultados(dfs)
    
    st.download_button(
        label="📥 Download Resultados (Excel)",
        data=buffer,
        file_name="resultados_simulacao.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


def criar_excel_resultados(dfs):
    """Cria arquivo Excel com os resultados"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for i, df in enumerate(dfs):
            df_to_save = df.copy()
            
            if 'Data' in df_to_save.columns:
                df_to_save.drop(columns=['Data'], inplace=True)
            
            if 'Demanda Atendida (m³/s)' in df_to_save.columns:
                df_to_save['Demanda Agregada (m³/s)'] = df_to_save['Demanda Atendida (m³/s)']
                segundos_mes = 2.592e6
                df_to_save['Demanda (hm³)'] = df_to_save['Demanda Atendida (m³/s)'] * (segundos_mes / 1e6)
            
            rename_map = {
                'Transferência Recebida (m³/s)': 'Transferência Recebida (m³/s)'
            }
            df_to_save.rename(columns=rename_map, inplace=True)
            
            cols_to_drop = ['Demanda Solicitada (m³/s)', 'Demanda Atendida (m³/s)']
            cols_existentes = [c for c in cols_to_drop if c in df_to_save.columns]
            if cols_existentes:
                df_to_save.drop(columns=cols_existentes, inplace=True)
            
            nome_aba = st.session_state.reservatorios_selecionados[i]['nome'][:30]
            df_to_save.to_excel(writer, sheet_name=nome_aba, index=False)
    
    output.seek(0)
    return output


def exibir_analise_garantia():
    """Exibe a análise de garantia e permanência"""
    resultados = st.session_state.resultados_simulacao
    dfs = resultados['dfs']
    params = resultados['params']
    modo = resultados['modo']
    vazao_conjunta = resultados['vazao_conjunta']
    
    total_meses = len(dfs[0])
    
    # ========================================================================
    # ANÁLISE DE VAZÕES TOTAIS DO SISTEMA
    # ========================================================================
    st.subheader("💧 Análise de Vazões Totais do Hidrossistema")
    
    # Calcular vazões totais agregadas
    vazoes_sistema = np.zeros(total_meses)
    for df in dfs:
        vazoes_sistema += df['Demanda Atendida (m³/s)'].to_numpy()
    
    # Calcular falhas do sistema
    falhas_conjuntas = np.zeros(total_meses, dtype=bool)
    if modo == "Paralelo":
        for df in dfs:
            mask_falha = (df['Falha'] == 'Sim').to_numpy()
            falhas_conjuntas = falhas_conjuntas | mask_falha
    else:
        falhas_conjuntas = np.ones(total_meses, dtype=bool)
        for df in dfs:
            mask_falha = (df['Falha'] == 'Sim').to_numpy()
            falhas_conjuntas = falhas_conjuntas & mask_falha
    
    # Criar DataFrame para análise
    df_vazoes = pd.DataFrame({
        'vazao': vazoes_sistema,
        'falha': falhas_conjuntas
    })
    df_vazoes['vazao_round'] = df_vazoes['vazao'].round(3)
    
    # Separar sucessos e falhas
    df_sucesso = df_vazoes[df_vazoes['falha'] == False]
    df_falha_sistema = df_vazoes[df_vazoes['falha'] == True]
    
    # Agrupar vazões de sucesso
    resumo_vazoes = df_sucesso.groupby('vazao_round').size().reset_index(name='permanencia')
    resumo_vazoes = resumo_vazoes.sort_values('vazao_round', ascending=False)
    resumo_vazoes['frequencia'] = (resumo_vazoes['permanencia'] / total_meses * 100).round(2)
    resumo_vazoes['garantia'] = resumo_vazoes['frequencia'].cumsum().round(2)
    
    # Adicionar falhas se houver
    count_falhas = len(df_falha_sistema)
    if count_falhas > 0:
        falha_row = pd.DataFrame({
            'vazao_round': ['FALHA'],
            'permanencia': [count_falhas],
            'frequencia': [(count_falhas / total_meses * 100)],
            'garantia': ['-']
        })
        resumo_vazoes = pd.concat([resumo_vazoes, falha_row], ignore_index=True)
    
    # Renomear colunas para exibição
    resumo_vazoes.columns = ['Vazão Total Sistema (m³/s)', 'Permanência (meses)', 'Frequência (%)', 'Garantia Acumulada (%)']
    
    # Exibir tabela
    st.dataframe(
        resumo_vazoes,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Vazão Total Sistema (m³/s)": st.column_config.NumberColumn(
                format="%.3f"
            ),
            "Permanência (meses)": st.column_config.NumberColumn(
                format="%d"
            ),
            "Frequência (%)": st.column_config.NumberColumn(
                format="%.2f%%"
            )
        }
    )
    
    # Estatísticas resumidas
    col1, col2, col3, col4 = st.columns(4)
    
    vazao_media = vazoes_sistema[~falhas_conjuntas].mean() if np.any(~falhas_conjuntas) else 0
    vazao_maxima = vazoes_sistema[~falhas_conjuntas].max() if np.any(~falhas_conjuntas) else 0
    vazao_minima = vazoes_sistema[~falhas_conjuntas].min() if np.any(~falhas_conjuntas) else 0
    num_falhas_sistema = np.sum(falhas_conjuntas)
    garantia_sistema = ((total_meses - num_falhas_sistema) / total_meses * 100)
    
    col1.metric("Vazão Média", f"{vazao_media:.3f} m³/s")
    col2.metric("Vazão Máxima", f"{vazao_maxima:.3f} m³/s")
    col3.metric("Vazão Mínima", f"{vazao_minima:.3f} m³/s")
    col4.metric("Garantia Sistema", f"{garantia_sistema:.2f}%")
    
    st.markdown("---")
    
    # ========================================================================
    # ANÁLISE DO SISTEMA (RESUMO)
    # ========================================================================
    st.subheader("📊 Resumo do Sistema")
    
    demanda_nominal_sistema = sum([p['demanda_nominal'] for p in params]) + vazao_conjunta
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Modo de Operação", modo)
    col2.metric("Meses Simulados", total_meses)
    col3.metric("Demanda Nominal Total (m³/s)", f"{demanda_nominal_sistema:.3f}")
    col4.metric("Meses com Falha", num_falhas_sistema)
    
    st.markdown("---")
    
    # ========================================================================
    # ANÁLISE INDIVIDUAL POR RESERVATÓRIO
    # ========================================================================
    st.subheader("📋 Detalhamento por Reservatório")
    
    for i, df in enumerate(dfs):
        nome = st.session_state.reservatorios_selecionados[i]['nome']
        demanda_nom = params[i]['demanda_nominal']
        
        with st.expander(f"🌊 {nome} (Demanda: {demanda_nom} m³/s)"):
            niveis_unicos = sorted(df['Racionamento (%)'].unique())
            garantia_acumulada = 0.0
            
            dados_tabela = []
            
            for rac in niveis_unicos:
                vazao_alvo = demanda_nom * (1 - rac / 100)
                df_filtro = df[(df['Racionamento (%)'] == rac) & (df['Falha'] == 'Não')]
                count = len(df_filtro)
                
                if count > 0:
                    nome_faixa = df_filtro['Modo Operação'].iloc[0]
                    freq = (count / total_meses) * 100
                    garantia_acumulada += freq
                    
                    dados_tabela.append({
                        'Nível de Seca': nome_faixa,
                        'Racionamento (%)': f"{rac:.1f}",
                        'Vazão (m³/s)': f"{vazao_alvo:.3f}",
                        'Permanência (meses)': count,
                        'Frequência (%)': f"{freq:.2f}",
                        'Garantia (%)': f"{garantia_acumulada:.2f}"
                    })
            
            count_falha = len(df[df['Falha'] == 'Sim'])
            if count_falha > 0:
                freq_falha = (count_falha / total_meses) * 100
                dados_tabela.append({
                    'Nível de Seca': '⚠️ FALHA',
                    'Racionamento (%)': 'FALHA',
                    'Vazão (m³/s)': '0.000',
                    'Permanência (meses)': count_falha,
                    'Frequência (%)': f"{freq_falha:.2f}",
                    'Garantia (%)': '-'
                })
            
            df_garantia = pd.DataFrame(dados_tabela)
            st.dataframe(df_garantia, use_container_width=True, hide_index=True)


if __name__ == "__main__":

    main()
